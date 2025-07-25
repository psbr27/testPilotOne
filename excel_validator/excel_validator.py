#!/usr/bin/env python3
import json
import re
import sys
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class ExcelValidator:
    def __init__(self, mapping_file: str):
        with open(mapping_file, "r") as f:
            self.config = json.load(f)

        self.wrap_text = self.config.get("wrap_text", False)
        self.urls_check = self.config.get("urls_check", {})
        self.methods_check = self.config.get("methods_check", {})
        self.pattern_match_check = self.config.get("pattern_match_check", {})

        self.issues = []

    def extract_curl_commands(self, df: pd.DataFrame) -> List[Dict]:
        """Extract curl commands from DataFrame"""
        curl_commands = []

        for index, row in df.iterrows():
            for col_name, cell_value in row.items():
                if pd.isna(cell_value):
                    continue

                cell_str = str(cell_value)
                if "curl" in cell_str.lower():
                    curl_commands.append(
                        {
                            "row": index,
                            "column": col_name,
                            "command": cell_str,
                            "original_value": cell_value,
                        }
                    )

        return curl_commands

    def parse_curl_command(self, command: str) -> Dict:
        """Parse curl command to extract URL, method, headers, and body"""
        parsed = {
            "url": None,
            "method": "GET",  # Default
            "headers": {},
            "body": None,
            "full_command": command,
        }

        # Extract URL
        url_match = re.search(r'"(https?://[^"]+)"', command)
        if url_match:
            parsed["url"] = url_match.group(1)

        # Extract method
        method_match = re.search(r"-X\s+(\w+)", command)
        if method_match:
            parsed["method"] = method_match.group(1)

        # Extract headers
        header_matches = re.findall(r'-H\s+[\'"]([^\'\"]+)[\'"]', command)
        for header in header_matches:
            if ":" in header:
                key, value = header.split(":", 1)
                parsed["headers"][key.strip()] = value.strip()

        # Extract body (data after -d or --data)
        body_match = re.search(r'(?:-d|--data)\s+[\'"]([^\'"]+)[\'"]', command)
        if body_match:
            parsed["body"] = body_match.group(1)

        return parsed

    def validate_url_mapping(self, parsed_curl: Dict) -> List[str]:
        """Validate URL against service mapping rules"""
        issues = []
        url = parsed_curl.get("url", "")

        if not url:
            return issues

        for url_pattern, expected_service in self.urls_check.items():
            if url_pattern in url:
                # Extract service from URL (between {})
                service_match = re.search(r"\{([^}]+)\}", url)
                if service_match:
                    actual_service = service_match.group(1)
                    if actual_service != expected_service:
                        issues.append(
                            f"URL service mismatch: found '{actual_service}', expected '{expected_service}' for pattern '{url_pattern}'"
                        )
                else:
                    issues.append(
                        f"No service placeholder found in URL for pattern '{url_pattern}'"
                    )

        return issues

    def validate_method_response(
        self, method: str, expected_status: str, curl_command: str = ""
    ) -> List[str]:
        """Validate HTTP method against expected response codes and headers"""
        issues = []
        method = method.upper() if method else ""

        if method in self.methods_check:
            method_config = self.methods_check[method]

            # Handle new generic format with "found" and "replace" properties
            if (
                isinstance(method_config, dict)
                and "found" in method_config
                and "replace" in method_config
            ):
                found_patterns = method_config["found"]
                replace_value = method_config["replace"]

                # Handle Content-Type validation (string pattern)
                if (
                    isinstance(found_patterns, str)
                    and "Content-Type:" in found_patterns
                ):
                    # This is a header validation (e.g., for PATCH)
                    expected_content_type = replace_value.split(":", 1)[1]
                    found_content_type = found_patterns.split(":", 1)[1]

                    # Check if curl command has the incorrect Content-Type header
                    if curl_command:
                        # Extract Content-Type from curl command headers
                        content_type_match = re.search(
                            r"-H\s+['\"]Content-Type:\s*([^'\"]+)['\"]",
                            curl_command,
                        )
                        if content_type_match:
                            actual_content_type = content_type_match.group(
                                1
                            ).strip()
                            if actual_content_type == found_content_type:
                                issues.append(
                                    f"Method '{method}' should use '{replace_value}', but found '{found_patterns}'"
                                )
                        else:
                            # If no Content-Type found, that's also an issue for methods that require it
                            issues.append(
                                f"Method '{method}' should have '{replace_value}' header, but no Content-Type found"
                            )

                # Handle status code validation (list of numbers)
                elif isinstance(found_patterns, list) and pd.notna(
                    expected_status
                ):
                    expected_status_str = str(expected_status).strip()

                    # Convert expected_status to comparable format
                    try:
                        status_value = float(expected_status_str)
                        if status_value in found_patterns:
                            issues.append(
                                f"Method '{method}' should return '{replace_value}', but Expected_Status shows specific code '{expected_status_str}'"
                            )
                    except (ValueError, TypeError):
                        # Handle string status codes
                        if expected_status_str in [
                            str(p) for p in found_patterns
                        ]:
                            issues.append(
                                f"Method '{method}' should return '{replace_value}', but Expected_Status shows specific code '{expected_status_str}'"
                            )

        return issues

    def validate_pattern_match(
        self, pattern_match_value: str
    ) -> Tuple[str, List[str]]:
        """Validate and fix Pattern_Match JSON content"""
        issues = []
        fixed_value = pattern_match_value

        if not self.pattern_match_check.get("json_fix", False):
            return fixed_value, issues

        if pd.isna(pattern_match_value):
            return fixed_value, issues

        pattern_str = str(pattern_match_value).strip()
        if not pattern_str:
            return fixed_value, issues

        # Check if it looks like JSON but isn't properly formatted
        if ("{" in pattern_str and "}" in pattern_str) or (
            pattern_str.startswith('"') and pattern_str.endswith('"')
        ):
            try:
                # Try to parse as JSON
                parsed = json.loads(pattern_str)
                # If successful, format properly
                fixed_value = json.dumps(parsed, indent=2)
                if fixed_value != pattern_str:
                    issues.append("Fixed JSON formatting in Pattern_Match")
            except json.JSONDecodeError:
                # Try to fix common JSON issues
                fixed_json = self.fix_common_json_issues(pattern_str)
                if fixed_json != pattern_str:
                    fixed_value = fixed_json
                    issues.append("Applied JSON fixes to Pattern_Match")

        return fixed_value, issues

    def validate_request_payload(self, payload_value: str) -> List[str]:
        """Validate Request_Payload references"""
        issues = []

        if pd.isna(payload_value):
            return issues

        payload_str = str(payload_value).strip()

        # Check if it's inline JSON that needs validation
        if (
            "{" in payload_str
            and "}" in payload_str
            and not payload_str.endswith(".json")
        ):
            # This is inline JSON - validate format
            try:
                json.loads(payload_str)
            except json.JSONDecodeError:
                issues.append(f"Invalid JSON format in Request_Payload")

        return issues

    def fix_common_json_issues(self, json_str: str) -> str:
        """Fix common JSON formatting issues"""
        # Remove extra spaces around colons and commas
        json_str = re.sub(r"\s*:\s*", ": ", json_str)
        json_str = re.sub(r"\s*,\s*", ", ", json_str)

        # Ensure proper quotes around keys and string values
        json_str = re.sub(r"(\w+):", r'"\1":', json_str)

        # Try to parse and format
        try:
            parsed = json.loads(json_str)
            return json.dumps(parsed, indent=2)
        except json.JSONDecodeError:
            return json_str

    def validate_excel_file(self, excel_file: str) -> Dict:
        """Main validation function"""
        results = {"issues": [], "fixes_applied": [], "sheets_processed": []}

        try:
            # Read Excel file
            xl_file = pd.ExcelFile(excel_file)

            for sheet_name in xl_file.sheet_names:
                results["sheets_processed"].append(sheet_name)
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                # Process each row systematically
                for index, row in df.iterrows():
                    # Extract relevant columns
                    command = row.get("Command", "")
                    expected_status = row.get("Expected_Status", "")
                    request_payload = row.get("Request_Payload", "")
                    pattern_match = row.get("Pattern_Match", "")

                    # Skip empty rows
                    if pd.isna(command) or not str(command).strip():
                        continue

                    # Parse curl command if present
                    if "curl" in str(command).lower():
                        parsed_curl = self.parse_curl_command(str(command))

                        # Validate URL mapping
                        url_issues = self.validate_url_mapping(parsed_curl)
                        for issue in url_issues:
                            results["issues"].append(
                                {
                                    "sheet": sheet_name,
                                    "row": index,
                                    "column": "Command",
                                    "type": "url_mapping",
                                    "issue": issue,
                                    "original_value": command,
                                }
                            )

                        # Validate method vs Expected_Status and headers
                        method_issues = self.validate_method_response(
                            parsed_curl.get("method", ""),
                            expected_status,
                            str(command),
                        )
                        for issue in method_issues:
                            # Determine which column the issue is about
                            if "Content-Type" in issue:
                                column = "Command"
                                original_value = command
                            else:
                                column = "Expected_Status"
                                original_value = expected_status

                            results["issues"].append(
                                {
                                    "sheet": sheet_name,
                                    "row": index,
                                    "column": column,
                                    "type": "method_response",
                                    "issue": issue,
                                    "original_value": original_value,
                                }
                            )

                    # Validate Request_Payload
                    if pd.notna(request_payload):
                        payload_issues = self.validate_request_payload(
                            request_payload
                        )
                        for issue in payload_issues:
                            results["issues"].append(
                                {
                                    "sheet": sheet_name,
                                    "row": index,
                                    "column": "Request_Payload",
                                    "type": "request_payload",
                                    "issue": issue,
                                    "original_value": request_payload,
                                }
                            )

                    # Validate Pattern_Match
                    if pd.notna(pattern_match):
                        fixed_pattern, pattern_issues = (
                            self.validate_pattern_match(pattern_match)
                        )
                        for issue in pattern_issues:
                            results["issues"].append(
                                {
                                    "sheet": sheet_name,
                                    "row": index,
                                    "column": "Pattern_Match",
                                    "type": "pattern_match",
                                    "issue": issue,
                                    "original_value": pattern_match,
                                    "fixed_value": fixed_pattern,
                                }
                            )

        except Exception as e:
            results["issues"].append(
                {
                    "sheet": "N/A",
                    "row": "N/A",
                    "column": "N/A",
                    "type": "file_error",
                    "issue": f"Error reading Excel file: {str(e)}",
                    "original_value": "N/A",
                }
            )

        return results

    def create_fixed_excel(
        self, original_file: str, output_file: str, validation_results: Dict
    ):
        """Create a fixed version of the Excel file"""
        try:
            workbook = load_workbook(original_file)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Apply text wrapping if enabled
                if self.wrap_text:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                cell.alignment = Alignment(wrap_text=True)

                # Apply fixes based on validation results
                for issue in validation_results["issues"]:
                    if issue["sheet"] == sheet_name:
                        row_idx = (
                            issue["row"] + 2
                        )  # pandas uses 0-based, openpyxl uses 1-based, +1 for header

                        if issue["type"] == "url_mapping":
                            # Fix URL service mapping in Command column
                            for col_idx, cell in enumerate(sheet[row_idx], 1):
                                if sheet.cell(1, col_idx).value == "Command":
                                    if (
                                        cell.value
                                        and "curl" in str(cell.value).lower()
                                    ):
                                        fixed_command = self.apply_url_fixes(
                                            str(cell.value), issue
                                        )
                                        cell.value = fixed_command
                                    break

                        elif issue["type"] == "method_response":
                            if "Content-Type" in issue["issue"]:
                                # Fix Content-Type in Command column
                                for col_idx, cell in enumerate(
                                    sheet[row_idx], 1
                                ):
                                    if (
                                        sheet.cell(1, col_idx).value
                                        == "Command"
                                    ):
                                        if (
                                            cell.value
                                            and "curl"
                                            in str(cell.value).lower()
                                        ):
                                            fixed_command = (
                                                self.apply_content_type_fixes(
                                                    str(cell.value), issue
                                                )
                                            )
                                            cell.value = fixed_command
                                        break
                            else:
                                # Fix Expected_Status column
                                for col_idx, cell in enumerate(
                                    sheet[row_idx], 1
                                ):
                                    if (
                                        sheet.cell(1, col_idx).value
                                        == "Expected_Status"
                                    ):
                                        if "specific code" in issue["issue"]:
                                            # Extract replacement value from issue message
                                            replace_match = re.search(
                                                r"should return '([^']+)'",
                                                issue["issue"],
                                            )
                                            if replace_match:
                                                cell.value = (
                                                    replace_match.group(1)
                                                )
                                            else:
                                                cell.value = "2xx"  # Fallback
                                        break

                        elif (
                            issue["type"] == "pattern_match"
                            and "fixed_value" in issue
                        ):
                            # Fix Pattern_Match JSON formatting
                            for col_idx, cell in enumerate(sheet[row_idx], 1):
                                if (
                                    sheet.cell(1, col_idx).value
                                    == "Pattern_Match"
                                ):
                                    cell.value = issue["fixed_value"]
                                    break

            workbook.save(output_file)
            return True

        except Exception as e:
            print(f"Error creating fixed Excel file: {str(e)}")
            return False

    def apply_url_fixes(self, command: str, issue: Dict) -> str:
        """Apply URL mapping fixes to curl commands"""
        fixed_command = command

        if "service mismatch" in issue["issue"]:
            # Extract expected service from issue description
            match = re.search(r"expected '([^']+)'", issue["issue"])
            if match:
                expected_service = match.group(1)
                # Replace the service in the URL
                fixed_command = re.sub(
                    r"\{[^}]+\}", f"{{{expected_service}}}", fixed_command
                )
        elif "No service placeholder found" in issue["issue"]:
            # Add service placeholder for localhost URLs
            if "127.0.0.1" in fixed_command or "localhost" in fixed_command:
                # Extract the URL pattern to determine correct service
                for url_pattern, expected_service in self.urls_check.items():
                    if url_pattern in fixed_command:
                        # Replace localhost with service placeholder
                        fixed_command = re.sub(
                            r"https?://127\.0\.0\.1:\d+",
                            f"http://{{{expected_service}}}:8081",
                            fixed_command,
                        )
                        break

        return fixed_command

    def apply_content_type_fixes(self, command: str, issue: Dict) -> str:
        """Apply Content-Type fixes to curl commands"""
        fixed_command = command

        if "Content-Type" in issue["issue"]:
            # Extract replacement Content-Type from the issue message
            if "should use" in issue["issue"]:
                # Extract the replacement pattern (e.g., "Content-Type:application/json-patch+json")
                match = re.search(r"should use '([^']+)'", issue["issue"])
                if match:
                    replacement_header = match.group(1)
                    expected_content_type = replacement_header.split(":", 1)[1]

                    # Check if Content-Type header exists and replace it
                    content_type_pattern = (
                        r"-H\s+['\"]Content-Type:\s*[^'\"]+['\"]"
                    )
                    if re.search(content_type_pattern, fixed_command):
                        # Replace existing Content-Type
                        fixed_command = re.sub(
                            content_type_pattern,
                            f"-H 'Content-Type:{expected_content_type}'",
                            fixed_command,
                        )
                    else:
                        # Add Content-Type header if missing
                        # Insert after the URL but before any data
                        url_match = re.search(r'"[^"]*"', fixed_command)
                        if url_match:
                            insert_pos = url_match.end()
                            before = fixed_command[:insert_pos]
                            after = fixed_command[insert_pos:]
                            fixed_command = f"{before} -H 'Content-Type:{expected_content_type}'{after}"

            elif (
                "should have" in issue["issue"] and "header" in issue["issue"]
            ):
                # Add missing Content-Type header
                match = re.search(
                    r"should have '([^']+)' header", issue["issue"]
                )
                if match:
                    replacement_header = match.group(1)
                    expected_content_type = replacement_header.split(":", 1)[1]

                    # Insert after the URL
                    url_match = re.search(r'"[^"]*"', fixed_command)
                    if url_match:
                        insert_pos = url_match.end()
                        before = fixed_command[:insert_pos]
                        after = fixed_command[insert_pos:]
                        fixed_command = f"{before} -H 'Content-Type:{expected_content_type}'{after}"

        return fixed_command


def main():
    if len(sys.argv) != 3:
        print("Usage: python excel_validator.py <excel_file> <mapping_file>")
        sys.exit(1)

    excel_file = sys.argv[1]
    mapping_file = sys.argv[2]

    validator = ExcelValidator(mapping_file)
    results = validator.validate_excel_file(excel_file)

    print("=== Validation Results ===")
    print(f"Sheets processed: {', '.join(results['sheets_processed'])}")
    print(f"Total issues found: {len(results['issues'])}")

    if results["issues"]:
        print("\n=== Issues Found ===")
        for i, issue in enumerate(results["issues"], 1):
            print(
                f"{i}. Sheet: {issue['sheet']}, Row: {issue['row']}, Column: {issue['column']}"
            )
            print(f"   Type: {issue['type']}")
            print(f"   Issue: {issue['issue']}")
            if "original_value" in issue:
                print(f"   Value: {str(issue['original_value'])[:100]}...")
            print()

        # Create fixed version
        output_file = excel_file.replace(".xlsx", "_fixed.xlsx")
        if validator.create_fixed_excel(excel_file, output_file, results):
            print(f"Fixed Excel file created: {output_file}")
        else:
            print("Failed to create fixed Excel file")
    else:
        print("No issues found!")


if __name__ == "__main__":
    main()
