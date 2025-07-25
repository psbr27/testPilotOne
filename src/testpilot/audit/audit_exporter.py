#!/usr/bin/env python3
"""
Audit Excel Exporter for TestPilot - Generates comprehensive audit reports
with properly formatted JSON data and detailed compliance information.
"""

import json
import os
from datetime import datetime
from typing import Any, Dict, List

# Import with fallback for dependencies
try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import logger with fallback
try:
    from ..utils.logger import get_logger

    logger = get_logger("TestPilot.AuditExporter")
except ImportError:
    import logging

    logger = logging.getLogger("TestPilot.AuditExporter")


class AuditExporter:
    """
    Excel exporter specifically designed for audit mode with enhanced
    formatting and comprehensive audit trail data.
    """

    def __init__(self):
        self.results_data = []
        self.summary_data = {}

    def export_audit_results(
        self,
        audit_results: List[Dict[str, Any]],
        audit_summary: Dict[str, Any],
        output_dir: str = "audit_reports",
        force_json: bool = False,
    ) -> str:
        """
        Export comprehensive audit results to Excel with multiple sheets
        and proper JSON formatting.

        Args:
            audit_results: List of detailed audit results
            audit_summary: Overall audit summary
            output_dir: Directory to save audit reports

        Returns:
            str: Path to generated Excel file
        """
        # Check dependencies or force JSON export
        if force_json or not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
            if force_json:
                logger.info("JSON export forced for testing purposes.")
            else:
                logger.warning(
                    "Excel export dependencies not available (pandas/openpyxl). Creating basic JSON report instead."
                )
            return self._export_json_fallback(
                audit_results, audit_summary, output_dir
            )

        # Create output directory
        os.makedirs(output_dir, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"audit_report_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Create workbook with multiple sheets
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Sheet 1: Detailed Audit Results
            self._export_detailed_results(audit_results, writer)

            # Sheet 2: Audit Summary
            self._export_audit_summary(audit_summary, writer)

            # Sheet 3: Compliance Report
            self._export_compliance_report(audit_results, writer)

            # Sheet 4: JSON Data (formatted)
            self._export_json_data(audit_results, writer)

            # Sheet 5: Differences Analysis
            self._export_differences_analysis(audit_results, writer)

        # Apply advanced formatting
        self._apply_advanced_formatting(filepath)

        logger.info(f"✅ Comprehensive audit report exported: {filepath}")
        return filepath

    def _export_detailed_results(self, audit_results: List[Dict], writer):
        """Export detailed audit results to main results sheet."""
        detailed_data = []

        for result in audit_results:
            # Format JSON data for display
            expected_json = self._format_json_for_excel(
                result.get("expected_pattern", "")
            )
            actual_json = self._format_json_for_excel(
                result.get("actual_response", "")
            )

            # Count differences
            differences_count = len(result.get("differences", []))
            http_errors_count = len(result.get("http_validation_errors", []))
            json_errors_count = len(result.get("json_validation_errors", []))

            # Create summary of issues
            issues_summary = []
            if http_errors_count > 0:
                issues_summary.append(
                    f"{http_errors_count} HTTP validation errors"
                )
            if json_errors_count > 0:
                issues_summary.append(
                    f"{json_errors_count} JSON validation errors"
                )
            if differences_count > 0:
                issues_summary.append(
                    f"{differences_count} pattern differences"
                )

            detailed_row = {
                "Test_Name": result.get("test_name", ""),
                "Timestamp": result.get("timestamp", ""),
                "Overall_Result": result.get("overall_result", ""),
                "HTTP_Method_Expected": result.get("http_method_expected", ""),
                "HTTP_Method_Actual": result.get("http_method_actual", ""),
                "Status_Code_Expected": result.get("status_code_expected", ""),
                "Status_Code_Actual": result.get("status_code_actual", ""),
                "Pattern_Match_Percentage": result.get("match_percentage", 0),
                "Differences_Count": differences_count,
                "HTTP_Validation_Errors": "; ".join(
                    result.get("http_validation_errors", [])
                ),
                "JSON_Validation_Errors": "; ".join(
                    result.get("json_validation_errors", [])
                ),
                "Issues_Summary": (
                    "; ".join(issues_summary)
                    if issues_summary
                    else "No issues"
                ),
                "Expected_Pattern": expected_json,
                "Actual_Response": actual_json,
                "Request_Details": json.dumps(
                    result.get("request_details", {}), indent=2
                ),
            }
            detailed_data.append(detailed_row)

        if PANDAS_AVAILABLE:
            df_detailed = pd.DataFrame(detailed_data)
            df_detailed.to_excel(
                writer, sheet_name="Detailed_Results", index=False
            )

    def _export_audit_summary(self, audit_summary: Dict, writer):
        """Export audit summary to dedicated sheet."""
        summary_data = [
            ["Metric", "Value"],
            ["Audit Mode", audit_summary.get("audit_mode", "")],
            ["Total Tests", audit_summary.get("total_tests", 0)],
            ["Passed Tests", audit_summary.get("passed_tests", 0)],
            ["Failed Tests", audit_summary.get("failed_tests", 0)],
            ["Error Tests", audit_summary.get("error_tests", 0)],
            ["Pass Rate (%)", f"{audit_summary.get('pass_rate', 0):.2f}%"],
            ["Compliance Status", audit_summary.get("compliance_status", "")],
            ["Generated At", audit_summary.get("generated_at", "")],
        ]

        df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
        df_summary.to_excel(writer, sheet_name="Audit_Summary", index=False)

    def _export_compliance_report(self, audit_results: List[Dict], writer):
        """Export compliance-focused report."""
        compliance_data = []

        for result in audit_results:
            is_compliant = result.get("overall_result") == "PASS"
            compliance_issues = []

            # Collect all compliance issues
            if result.get("http_validation_errors"):
                compliance_issues.extend(result["http_validation_errors"])
            if result.get("json_validation_errors"):
                compliance_issues.extend(result["json_validation_errors"])
            if result.get("differences"):
                for diff in result["differences"]:
                    compliance_issues.append(
                        f"{diff['type']} at {diff['field_path']}: expected '{diff['expected_value']}', got '{diff['actual_value']}'"
                    )

            compliance_row = {
                "Test_Name": result.get("test_name", ""),
                "Compliant": "YES" if is_compliant else "NO",
                "Compliance_Issues_Count": len(compliance_issues),
                "Compliance_Issues": (
                    "; ".join(compliance_issues)
                    if compliance_issues
                    else "None"
                ),
                "Risk_Level": self._assess_risk_level(result),
                "Remediation_Required": "NO" if is_compliant else "YES",
            }
            compliance_data.append(compliance_row)

        df_compliance = pd.DataFrame(compliance_data)
        df_compliance.to_excel(
            writer, sheet_name="Compliance_Report", index=False
        )

    def _export_json_data(self, audit_results: List[Dict], writer):
        """Export properly formatted JSON data for technical analysis."""
        json_data = []

        for result in audit_results:
            json_row = {
                "Test_Name": result.get("test_name", ""),
                "Expected_JSON": self._format_json_for_excel(
                    result.get("expected_pattern", "")
                ),
                "Actual_JSON": self._format_json_for_excel(
                    result.get("actual_response", "")
                ),
                "JSON_Valid": (
                    "YES" if not result.get("json_validation_errors") else "NO"
                ),
                "Pattern_Match": (
                    "100%"
                    if result.get("overall_result") == "PASS"
                    else "FAILED"
                ),
            }
            json_data.append(json_row)

        df_json = pd.DataFrame(json_data)
        df_json.to_excel(writer, sheet_name="JSON_Data", index=False)

    def _export_differences_analysis(self, audit_results: List[Dict], writer):
        """Export detailed differences analysis."""
        differences_data = []

        for result in audit_results:
            test_name = result.get("test_name", "")
            for diff in result.get("differences", []):
                diff_row = {
                    "Test_Name": test_name,
                    "Difference_Type": diff.get("type", ""),
                    "Field_Path": diff.get("field_path", ""),
                    "Expected_Value": str(diff.get("expected_value", "")),
                    "Actual_Value": str(diff.get("actual_value", "")),
                    "Impact": self._assess_difference_impact(diff),
                }
                differences_data.append(diff_row)

        if differences_data:
            df_differences = pd.DataFrame(differences_data)
            df_differences.to_excel(
                writer, sheet_name="Differences_Analysis", index=False
            )
        else:
            # Create empty sheet with headers
            df_empty = pd.DataFrame(
                columns=[
                    "Test_Name",
                    "Difference_Type",
                    "Field_Path",
                    "Expected_Value",
                    "Actual_Value",
                    "Impact",
                ]
            )
            df_empty.to_excel(
                writer, sheet_name="Differences_Analysis", index=False
            )

    def _format_json_for_excel(self, json_data: str) -> str:
        """Format JSON data for proper display in Excel."""
        if not json_data:
            return ""

        try:
            if isinstance(json_data, str):
                parsed = json.loads(json_data)
            else:
                parsed = json_data
            return json.dumps(parsed, indent=2, ensure_ascii=False)
        except (json.JSONDecodeError, TypeError):
            return str(json_data)

    def _assess_risk_level(self, result: Dict) -> str:
        """Assess risk level based on audit result."""
        if result.get("overall_result") == "PASS":
            return "LOW"

        errors_count = (
            len(result.get("http_validation_errors", []))
            + len(result.get("json_validation_errors", []))
            + len(result.get("differences", []))
        )

        if errors_count >= 5:
            return "HIGH"
        elif errors_count >= 2:
            return "MEDIUM"
        else:
            return "LOW"

    def _assess_difference_impact(self, diff: Dict) -> str:
        """Assess the impact level of a specific difference."""
        diff_type = diff.get("type", "").lower()
        field_path = diff.get("field_path", "").lower()

        # Critical fields that have high impact
        critical_fields = ["status", "error", "result", "success", "code"]
        if any(field in field_path for field in critical_fields):
            return "HIGH"

        if diff_type == "missing":
            return "HIGH"
        elif diff_type == "mismatch":
            return "MEDIUM"
        else:
            return "LOW"

    def _apply_advanced_formatting(self, filepath: str):
        """Apply advanced Excel formatting to the audit report."""
        try:
            from openpyxl.styles import Border, Side

            workbook = load_workbook(filepath)

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            pass_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            fail_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]

                # Apply header formatting
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                # Apply conditional formatting for results
                if sheet_name == "Detailed_Results":
                    result_col = None
                    for idx, cell in enumerate(ws[1], 1):
                        if cell.value == "Overall_Result":
                            result_col = idx
                            break

                    if result_col:
                        for row in ws.iter_rows(
                            min_row=2, min_col=result_col, max_col=result_col
                        ):
                            for cell in row:
                                if cell.value == "PASS":
                                    cell.fill = pass_fill
                                elif cell.value == "FAIL":
                                    cell.fill = fail_fill

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(
                        max_length + 2, 100
                    )  # Cap at 100 characters
                    ws.column_dimensions[column_letter].width = adjusted_width

            workbook.save(filepath)
            logger.debug(f"Advanced formatting applied to {filepath}")

        except Exception as e:
            logger.warning(f"Could not apply advanced formatting: {e}")

    def _export_json_fallback(
        self,
        audit_results: List[Dict[str, Any]],
        audit_summary: Dict[str, Any],
        output_dir: str,
    ) -> str:
        """Export audit results as JSON when Excel dependencies are not available"""
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"audit_report_{timestamp}.json"
        filepath = os.path.join(output_dir, filename)

        report_data = {
            "audit_summary": audit_summary,
            "audit_results": audit_results,
            "export_format": "JSON (Excel dependencies not available)",
            "generated_at": datetime.now().isoformat(),
        }

        with open(filepath, "w") as f:
            json.dump(report_data, f, indent=2)

        logger.info(f"JSON audit report exported: {filepath}")
        return filepath
